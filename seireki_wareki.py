import openpyxl as excel

wareki_tabel = [{"name": "明治", "start": 1868, "end": 1912},
                {"name": "大正", "start": 1912, "end": 1926},
                {"name": "昭和", "start": 1926, "end": 1989},
                {"name": "平成", "start": 1989, "end": 2019},
                {"name": "令和", "start": 2019, "end": 9999}]


def seireki_wareki(year):
    for w in wareki_tabel:
        if w["start"] <= year < w["end"]:
            y = str(year - w["start"] + 1) + "年"
            if y == "1": y = "元年"
            return w["name"] + y

    return "不明"


book = excel.Workbook()
sheet = book.active

sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

start_y = 1930

for i in range(100):
    sei = start_y + i
    wa = seireki_wareki(sei)
    sheet.cell(row=i + 1, column=1, value=f"{sei}年")
    sheet.cell(row=i + 1, column=2, value=wa)

book.save('seireki_wareki.xlsx')
